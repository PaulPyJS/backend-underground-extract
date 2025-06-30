import pdfplumber
import logging
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import tkinter as tk
from tkinter import Tk, filedialog, messagebox, ttk
import re

import statistics
from decimal import Decimal, getcontext
getcontext().prec = 10

logging.getLogger("pdfminer").setLevel(logging.ERROR)

# === Script : EXTRACT VALUE UNDER KEYWORD IN A TABLE-LIKE FORMAT - ESSAIS PRESSIOMETRIQUES ===
# = v2 : Debug by highlight of keyword in the page : still only one page at a time
# = v3 : Adding UI to validate data from extract to manually adapt if necessary
# = v4 : Adding Depth from user to create list based on a STEP, DEPTH START/END
# = v5 : Adapting code to detect variation between Y coordinate from the values = No values or issues
#        Allowing user to visualize these variations within UI
# = v6 : Adding multi pages processing based on borehole names + multi pages UI
# = v6.5 : Add log on UI and asking user for the keywords
# = v7 : Exporting the data in Excel EM|Pl for each borehole after last validate on the verif UI + log on UI
# = v7.5 : Re-opening data after validated lists is possible
#

def detect_y_anomalies(y_val_list, keyword):
    if len(y_val_list) < 3:
        return [v for _, v in y_val_list], []

    y_val_list = sorted(y_val_list, key=lambda x: x[0])
    y_positions = [y for y, _ in y_val_list]
    dy_list = [y2 - y1 for y1, y2 in zip(y_positions, y_positions[1:])]

    median_dy = statistics.median(dy_list)
    min_dy = 0.7 * median_dy
    max_dy = 1.3 * median_dy

    print(f"\n📏 Médiane des écarts Y pour '{keyword}': {median_dy:.2f} pts")
    print(f"🔍 Seuils → trop petit: < {min_dy:.2f} pts | trop grand: > {max_dy:.2f} pts")
    print(f"↕️ Écarts Y pour '{keyword}':")

    output = []
    logs = []

    highlight_indices = []

    for i in range(len(y_val_list) - 1):
        y1, v1 = y_val_list[i]
        y2, v2 = y_val_list[i + 1]
        dy = abs(y2 - y1)

        print(f"  dy[{i}] = {dy:.2f} pts entre {v1} et {v2}")

        output.append(v1)

        dy = Decimal(str(y2 - y1))
        median_dy = Decimal(str(statistics.median(dy_list)))
        min_dy = median_dy * Decimal("0.7")
        max_dy = median_dy * Decimal("1.3")

        if dy > max_dy:
            print("    → TROU détecté")
            output.append(None)
            logs.append(f" NULL : Trou détecté pour '{keyword}' entre {v1} et {v2} (écart Y = {dy:.1f} pts)")
        elif dy < min_dy:
            print("    → TROP PETIT")
            highlight_indices.append(len(output) - 1)
            highlight_indices.append(len(output))
            logs.append(f" ⚠️  : Espacement trop petit pour '{keyword}' entre {v1} et {v2} (écart Y = {dy:.1f} pts)")

    output.append(y_val_list[-1][1])  # Dernière valeur

    return output, logs, highlight_indices


def detect_sondage_name(words):
    pattern = re.compile(r"\bSP\d{1,4}\b")
    for w in words:
        text = w.get('text', '')
        if pattern.fullmatch(text.strip()):
            return text.strip()
    return None



class PDFKeywordExtractor:
    def __init__(self, pdf_path, keywords, dpi=150, tolerances=None, column_distance_threshold=15):
        self.pdf_path = pdf_path
        self.keywords = keywords
        self.dpi = dpi
        self.column_distance_threshold = column_distance_threshold
        self.drag_data = {}

        # Tolerance par mot-clef
        self.tolerances = tolerances or {
            self.keywords[0]: {"left": 10, "right": 30, "min_dy": 50},
            self.keywords[1]: {"left": 10, "right": 30, "min_dy": 50},
            self.keywords[2]: {"left": 10, "right": 54, "min_dy": 50}
        }


    def pt_to_px(self, val):
        return val * self.dpi / 72

    def get_keyword_x_positions(self, words):
        positions = {}
        for kw in self.keywords:
            for w in words:
                if w['text'].strip().lower() == kw.lower():
                    x = (w['x0'] + w['x1']) / 2
                    positions[kw] = x
                    break
        return positions

    # =========================== DEBUG =======================================
    #
    def highlight_keywords_on_page(self, page):
        words = page.extract_words()
        im = page.to_image(resolution=self.dpi)
        pil_image = im.original

        fig, ax = plt.subplots()
        ax.imshow(pil_image, origin="upper")


        print("\n--- Coordonnées des mots-clés trouvés (en pts) ---")
        found_any = False
        for word in words:
            text = word.get('text', '')
            for kw in self.keywords:
                if text.strip().lower() == kw.lower():
                    x0 = self.pt_to_px(word['x0'])
                    x1 = self.pt_to_px(word['x1'])
                    width = x1 - x0

                    top = self.pt_to_px(word['top'])
                    height = self.pt_to_px(word['bottom'] - word['top'])

                    rect = Rectangle((x0, top), width, height, edgecolor='red', linewidth=2, facecolor='none')
                    ax.add_patch(rect)

                    center_x = x0 + width / 2
                    center_y = top + height / 2
                    ax.plot(center_x, center_y, 'ro', markersize=3)

                    print(f"{kw:<10} → x={word['x0']:.1f}, y={word['top']:.1f}")
                    found_any = True
                    break

        if not found_any:
            print("❗ Aucun mot-clé trouvé pour surlignage.")

        ax.set_title("Mots-clés surlignés")
        ax.axis("off")
        plt.tight_layout()
        plt.show()
    #
    # =========================== DEBUG =======================================

    # = UI for validation of the list, and suppression of strange values
    #
    def show_validation_ui_with_tabs(self, results_by_sondage):
        root = tk.Tk()
        root.title("Validation des sondages")
        notebook = ttk.Notebook(root)

        main_container = tk.Frame(root)
        main_container.pack(fill='both', expand=True)

        notebook = ttk.Notebook(main_container)
        notebook.pack(side=tk.LEFT, fill='both', expand=True)

        right_panel = tk.Frame(main_container)
        right_panel.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=10)

        tk.Label(right_panel, text="✅ Sondages validés", font=("Helvetica", 10, "bold")).pack(pady=(0, 5))

        validated_listbox = tk.Listbox(right_panel, height=15, width=25)
        validated_listbox.pack()

        self.drag_data = {}
        sondage_tabs = {}
        validated_sondages = {}

        def generate_canvas_with_guidelines(parent, height, width, line_height=16, bg="#f9f9f9"):
            canvas = tk.Canvas(parent, width=width, height=height, bg=bg, highlightthickness=0)
            for y in range(0, height, line_height):
                canvas.create_line(0, y, width, y, fill="#dddddd")
            return canvas

        def update_listbox(lb, data, red_idx=None):
            red_idx = red_idx or []
            lb.delete(0, tk.END)
            for i, val in enumerate(data):
                display = "NULL" if val is None else val
                lb.insert(tk.END, display)
                if val is None:
                    lb.itemconfig(i, {'fg': '#aa0000'})
                if i in red_idx:
                    lb.itemconfig(i, {'fg': 'red'})

        def add_value(list_ref, listbox):
            popup = tk.Toplevel()
            popup.title("Ajouter une valeur")
            tk.Label(popup, text="Nouvelle valeur :").pack(padx=10, pady=5)
            entry = tk.Entry(popup)
            entry.pack(padx=10, pady=5)

            def submit():
                try:
                    val = float(entry.get())
                    selection = listbox.curselection()
                    if selection:
                        index = selection[0] + 1
                        list_ref.insert(index, val)
                    else:
                        list_ref.append(val)
                    update_listbox(listbox, list_ref)
                    popup.destroy()
                except ValueError:
                    messagebox.showwarning("Erreur", "Veuillez entrer un nombre valide.")

            tk.Button(popup, text="Valider", command=submit).pack(pady=10)
            entry.focus()

        def remove_selected(list_ref, listbox):
            selection = listbox.curselection()
            if selection:
                index = selection[0]
                del list_ref[index]
                update_listbox(listbox, list_ref)

        def set_null(list_ref, listbox):
            selection = listbox.curselection()
            if selection:
                index = selection[0]
                list_ref[index] = None
                update_listbox(listbox, list_ref)

        def validate_tab(sondage_name):
            tab = sondage_tabs[sondage_name]
            print(f"\n✅ Données validées pour sondage {sondage_name}")
            for i in range(len(tab["Depth"])):
                row = [tab[k][i] for k in ("Pf*", "Pl*", "Module", "Depth")]
                print(row)
            validated_sondages[sondage_name] = {
                "Depth": tab["Depth"],
                "Pf*": tab["Pf*"],
                "Pl*": tab["Pl*"],
                "Module": tab["Module"]
            }
            validated_listbox.insert(tk.END, sondage_name)
            notebook.forget(tab["frame"])
            del sondage_tabs[sondage_name]
            if not sondage_tabs:
                messagebox.showinfo("Terminé",
                                    "Tous les sondages ont été validés.\nVous pouvez maintenant exporter les données.")

        def reopen_selected_tab():
            selection = validated_listbox.curselection()
            if not selection:
                return

            sondage_name = validated_listbox.get(selection[0])
            data = validated_sondages.pop(sondage_name)
            validated_listbox.delete(selection[0])

            frame = ttk.Frame(notebook)
            notebook.add(frame, text=sondage_name)
            sondage_tabs[sondage_name] = {"frame": frame}

            tk.Label(frame, text="UI de vérification", font=("Helvetica", 11, "bold")).pack(pady=(10, 5))

            header_frame = tk.Frame(frame)
            header_frame.pack()
            for label in ["Profondeur", self.keywords[0], self.keywords[1], self.keywords[2]]:
                tk.Label(header_frame, text=label, font=("Helvetica", 9, "normal"), width=10).pack(side=tk.LEFT, padx=8)

            main_frame = tk.Frame(frame)
            main_frame.pack(padx=10, pady=5)

            columns = ["Depth", "Pf*", "Pl*", "Module"]
            for col, key in enumerate(columns):
                col_frame = tk.Frame(main_frame)
                col_frame.grid(row=1, column=col, padx=10)

                canvas = generate_canvas_with_guidelines(col_frame, height=360, width=110)
                canvas.pack()
                listbox = tk.Listbox(canvas, height=22, width=10, font=("Helvetica", 10), bd=0, highlightthickness=0)
                canvas.create_window((55, 180), window=listbox)

                sondage_tabs[sondage_name][key] = data[key]

                update_listbox(listbox, data[key])

                btn_frame = tk.Frame(col_frame)
                btn_frame.pack(pady=(6, 0))

                btn_style = {"font": ("Helvetica", 10, "bold"), "width": 2}
                tk.Button(btn_frame, text="+", **btn_style,
                          command=lambda d=data[key], lb=listbox: add_value(d, lb)).pack(side=tk.LEFT, padx=2)
                tk.Button(btn_frame, text="-", **btn_style,
                          command=lambda d=data[key], lb=listbox: remove_selected(d, lb)).pack(side=tk.LEFT, padx=2)

                tk.Button(col_frame, text="NULL", width=6, relief="groove", font=("Helvetica", 9),
                          command=lambda d=data[key], lb=listbox: set_null(d, lb)).pack(pady=(4, 2))

                bind_drag_events(listbox, data[key], col)

            tk.Button(frame, text="Valider", font=("Helvetica", 10, "bold"),
                      command=lambda s=sondage_name: validate_tab(s)).pack(pady=10)

        def export_to_excel():
            import openpyxl
            from openpyxl.styles import Font

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sondages"

            col_index = 1
            for sondage_name, data in validated_sondages.items():
                # Sondage name, line 1
                ws.cell(row=1, column=col_index, value=sondage_name).font = Font(bold=True)

                # Column header, line 2
                headers = ["Profondeur", "EM", "PL"]
                for offset, header in enumerate(headers):
                    ws.cell(row=2, column=col_index + offset, value=header).font = Font(italic=True)

                # Données à partir de la ligne 3
                for i, val in enumerate(data["Depth"]):
                    ws.cell(row=i + 3, column=col_index, value=val)
                for i, val in enumerate(data["Module"]):
                    ws.cell(row=i + 3, column=col_index + 1, value=val)
                for i, val in enumerate(data["Pl*"]):
                    ws.cell(row=i + 3, column=col_index + 2, value=val)

                col_index += 4  # 3 colonnes + 1 vide

            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Enregistrer sous"
            )
            if save_path:
                wb.save(save_path)
                messagebox.showinfo("Export réussi", f"Fichier enregistré :\n{save_path}")

        tk.Button(right_panel, text="Exporter vers Excel", font=("Helvetica", 10, "bold"),
                  command=export_to_excel).pack(pady=20)

        tk.Button(right_panel, text="🔁 Rouvrir un sondage", font=("Helvetica", 10),
                  command=reopen_selected_tab).pack(pady=5)



        for sondage, data in results_by_sondage.items():
            frame = ttk.Frame(notebook)
            notebook.add(frame, text=sondage)
            sondage_tabs[sondage] = {"frame": frame}

            tk.Label(frame, text="UI de vérification", font=("Helvetica", 11, "bold")).pack(pady=(10, 5))

            header_frame = tk.Frame(frame)
            header_frame.pack()
            for label in ["Profondeur", self.keywords[0], self.keywords[1], self.keywords[2]]:
                tk.Label(header_frame, text=label, font=("Helvetica", 9, "normal"), width=10).pack(side=tk.LEFT, padx=8)

            main_frame = tk.Frame(frame)
            main_frame.pack(padx=10, pady=5)

            columns = ["Depth", "Pf*", "Pl*", "Module"]
            listboxes = []
            for col, key in enumerate(columns):
                col_frame = tk.Frame(main_frame)
                col_frame.grid(row=1, column=col, padx=10)

                # Génère un canvas avec lignes horizontales
                canvas = generate_canvas_with_guidelines(col_frame, height=360, width=110)
                canvas.pack()

                # Crée la listbox par-dessus
                listbox = tk.Listbox(canvas, height=22, width=10, font=("Helvetica", 10), bd=0, highlightthickness=0)
                canvas.create_window((55, 180), window=listbox)

                # 🔁 Important : copie locale pour éviter les bugs d'index sur les couleurs
                local_data = list(data[key])
                red = list(data["RedFlags"].get(key, [])) if key != "Depth" else []
                update_listbox(listbox, local_data, red)
                sondage_tabs[sondage][key] = local_data
                listboxes.append(listbox)  # pour la scrollbar partagée

                btn_frame = tk.Frame(col_frame)
                btn_frame.pack(pady=(6, 0))

                btn_style = {"font": ("Helvetica", 10, "bold"), "width": 2}
                tk.Button(btn_frame, text="+", **btn_style,
                          command=lambda d=data[key], lb=listbox: add_value(d, lb)).pack(side=tk.LEFT, padx=2)
                tk.Button(btn_frame, text="-", **btn_style,
                          command=lambda d=data[key], lb=listbox: remove_selected(d, lb)).pack(side=tk.LEFT, padx=2)

                tk.Button(col_frame, text="NULL", width=6, relief="groove", font=("Helvetica", 9),
                          command=lambda d=data[key], lb=listbox: set_null(d, lb)).pack(pady=(4, 2))

                def on_drag_start(event, idx=col):
                    self.drag_data[idx] = listbox.nearest(event.y)

                def on_drag_motion(event):
                    pass

                def bind_drag_events(lb, data_ref, idx):
                    def on_drag_start(event):
                        self.drag_data[idx] = lb.nearest(event.y)

                    def on_drag_motion(event):
                        pass

                    def on_drag_release(event):
                        from_idx = self.drag_data.get(idx)
                        if from_idx is None:
                            return
                        to_idx = lb.nearest(event.y)
                        if from_idx != to_idx:
                            item = data_ref.pop(from_idx)
                            data_ref.insert(to_idx, item)
                            update_listbox(lb, data_ref)

                    lb.bind("<ButtonPress-1>", on_drag_start)
                    lb.bind("<B1-Motion>", on_drag_motion)
                    lb.bind("<ButtonRelease-1>", on_drag_release)

                bind_drag_events(listbox, data[key], col)

            tk.Button(frame, text="Valider", font=("Helvetica", 10, "bold"),
                      command=lambda s=sondage: validate_tab(s)).pack(pady=10)

        root.mainloop()

    def ask_user_for_depth_range(self, sondage_name):
        depth_values = []

        def on_submit():
            try:
                d_start = float(entry_start.get())
                d_end = float(entry_end.get())
                d_step = float(entry_step.get())

                if d_start >= d_end or d_step <= 0:
                    raise ValueError

                for i in range(int((d_end - d_start) / d_step) + 1):
                    depth_values.append(round(d_start + i * d_step, 3))

                win.destroy()
            except ValueError:
                messagebox.showerror("Erreur", "Veuillez entrer des valeurs valides.")

        win = tk.Toplevel()
        win.title(f"Paramètres du sondage - {sondage_name}")
        win.attributes("-topmost", True)
        win.lift()

        tk.Label(win, text=f"Sondage : {sondage_name}", font=("Helvetica", 11, "bold")).pack(pady=(10, 5))

        tk.Label(win, text="Profondeur de départ :").pack()
        entry_start = tk.Entry(win)
        entry_start.pack()

        tk.Label(win, text="Profondeur de fin :").pack()
        entry_end = tk.Entry(win)
        entry_end.pack()

        tk.Label(win, text="Pas (ex: 0.2) :").pack()
        entry_step = tk.Entry(win)
        entry_step.pack()

        tk.Button(win, text="Valider", command=on_submit).pack(pady=10)

        win.grab_set()
        win.wait_window()

        return depth_values



    def extract_values_near_keyword(self, words, keyword):
        ref_word = next((w for w in words if w['text'].strip().lower() == keyword.lower()), None)
        if not ref_word:
            return []

        tol = self.tolerances.get(keyword, {
            "left": 10,
            "right": 30,
            "min_dy": 50
        })

        x_ref = (ref_word['x0'] + ref_word['x1']) / 2
        y_ref = ref_word['top']

        values = []
        for w in words:
            try:
                val = float(w['text'].replace(",", "."))
            except ValueError:
                continue

            x_c = (w['x0'] + w['x1']) / 2
            y_c = w['top']

            if (x_ref - tol['left'] <= x_c <= x_ref + tol['right']) and (y_c > y_ref + tol['min_dy']):
                values.append((y_c, val))

        return sorted(values, key=lambda x: x[0])





    # = PROCESSING ALL THE PAGES FOR A PDF DOCUMENT
    def process_all_pages(self):
        results_by_sondage = {}
        depths_by_sondage = {}

        with pdfplumber.open(self.pdf_path) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                print(f"\n📄 Traitement page {page_idx + 1}")
                words = page.extract_words()
                sondage_name = detect_sondage_name(words) or f"Page {page_idx + 1}"

                # Extraction des valeurs par mot-clé
                values_by_keyword = {}
                for kw in self.keywords:
                    values = self.extract_values_near_keyword(words, kw)
                    values_by_keyword[kw] = values
                    print(f"🔍 {kw} : {len(values)} valeurs")

                # Vérifie la position de Pf et Pl
                x_positions = self.get_keyword_x_positions(words)
                is_combined = False
                if self.keywords[0] in x_positions and self.keywords[1] in x_positions:
                    distance = abs(x_positions[self.keywords[0]] - x_positions[self.keywords[1]])
                    if distance <= self.column_distance_threshold:
                        is_combined = True

                if is_combined:
                    print("✅ Pf et Pl combinés → traitement spécial")
                    pf_values = values_by_keyword[self.keywords[0]]
                    pl_values = values_by_keyword[self.keywords[1]]
                    if pf_values != pl_values or len(pf_values) % 2 != 0:
                        print("⚠️ Données incohérentes pour traitement spécial. Saut de cette page.")
                        continue

                    pf_final, pl_final = [], []
                    for i in range(len(pf_values) - 2, -1, -2):
                        a, b = pf_values[i][1], pf_values[i + 1][1]
                        if a < b:
                            pf_final.append((pf_values[i][0], a))
                            pl_final.append((pf_values[i + 1][0], b))
                        else:
                            pf_final.append((pf_values[i + 1][0], b))
                            pl_final.append((pf_values[i][0], a))

                    pf_final.reverse()
                    pl_final.reverse()
                else:
                    print("✅ Pf et Pl séparés → traitement standard")
                    pf_final = values_by_keyword[self.keywords[0]]
                    pl_final = values_by_keyword[self.keywords[1]]

                em_values = values_by_keyword[self.keywords[2]]

                # Analyse des anomalies Y
                pf_list, pf_logs, pf_red = detect_y_anomalies(pf_final, self.keywords[0])
                pl_list, pl_logs, pl_red = detect_y_anomalies(pl_final, self.keywords[1])
                em_list, em_logs, em_red = detect_y_anomalies(em_values, self.keywords[2])

                # Une seule demande de profondeur par sondage
                if sondage_name not in depths_by_sondage:
                    depths_by_sondage[sondage_name] = self.ask_user_for_depth_range(sondage_name)
                depths = depths_by_sondage[sondage_name]

                # Ajout ou accumulation des données par sondage
                if sondage_name not in results_by_sondage:
                    results_by_sondage[sondage_name] = {
                        "Pf*": pf_list,
                        "Pl*": pl_list,
                        "Module": em_list,
                        "Depth": depths,  # ✅ Ajouté une seule fois ici
                        "RedFlags": {
                            "Pf*": pf_red,
                            "Pl*": pl_red,
                            "Module": em_red
                        }
                    }
                else:
                    results_by_sondage[sondage_name]["Pf*"] += pf_list
                    results_by_sondage[sondage_name]["Pl*"] += pl_list
                    results_by_sondage[sondage_name]["Module"] += em_list
                    results_by_sondage[sondage_name]["RedFlags"]["Pf*"] += pf_red
                    results_by_sondage[sondage_name]["RedFlags"]["Pl*"] += pl_red
                    results_by_sondage[sondage_name]["RedFlags"]["Module"] += em_red

        self.show_validation_ui_with_tabs(results_by_sondage)



# === Lancement ===
def choose_pdf():
    root = Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Choisir un fichier PDF",
        filetypes=[("Fichiers PDF", "*.pdf")]
    )


if __name__ == "__main__":
    pdf_path = choose_pdf()
    if pdf_path:
        extractor = PDFKeywordExtractor(
            pdf_path,
            keywords=["Pf*", "Pl*", "Module"],
        )
        extractor.process_all_pages()
    else:
        print("Aucun fichier sélectionné.")
